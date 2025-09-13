# core/views.py
from django.shortcuts import render
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, Http404, StreamingHttpResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Max
from django.db.models import F, Q, Value, FloatField, IntegerField, Case, When, CharField, Func
from django.db.models.functions import Lower, Greatest
try:
    from django.contrib.postgres.search import TrigramSimilarity, TrigramWordSimilarity  # type: ignore
except Exception:  # pragma: no cover
    from django.contrib.postgres.search import TrigramSimilarity  # type: ignore
    TrigramWordSimilarity = None  # type: ignore
from django.utils.timezone import localtime
# NOTE: we rely on PostgreSQL unaccent(), not on unidecode
from django.shortcuts import render




# Modele – fără dubluri
from .models import (
    Reserve, Association, ReserveAssociationYear,
    Occurrence, Species, SiteHabitat, Site, Habitat,
)

# CSV/XLSX
import csv
from io import StringIO, BytesIO
import unicodedata
import difflib
from django.utils.html import strip_tags
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.http import require_GET

class _Echo:
    def write(self, value):  # csv.writer cere un .write()
        return value

def _stream_csv(filename: str, headers, rows_iter):
    """rows_iter -> iterabil de liste/tupluri; folosește csv.writer streaming."""
    pseudo = _Echo()
    writer = csv.writer(pseudo)

    def generator():
        yield writer.writerow(headers)
        for row in rows_iter:
            yield writer.writerow(row)

    resp = StreamingHttpResponse(generator(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def _paginate(request, qs, default=50, max_per_page=200):
    """Returnează (paginator, page_obj) pe baza ?page=&per_page=."""
    try:
        per_page = min(max_per_page, max(1, int(request.GET.get("per_page", default))))
    except ValueError:
        per_page = default
    try:
        page_no = int(request.GET.get("page", 1))
    except ValueError:
        page_no = 1
    paginator = Paginator(qs, per_page)
    return paginator, paginator.get_page(page_no)


def _resolve_reserve(q):
    """
    Permite căutare fie după id (număr), fie după nume (case-insensitive contains).
    Returnează un obiect Reserve sau ridică Http404.
    """
    if not q:
        raise Http404("Lipsește parametrul 'reserve'")
    if str(q).isdigit():
        try:
            return Reserve.objects.get(pk=int(q))
        except Reserve.DoesNotExist:
            raise Http404("Rezervație inexistentă")
    # nume parțial
    try:
        return Reserve.objects.get(name__iexact=q)
    except Reserve.DoesNotExist:
        # fallback: primul match parțial
        match = Reserve.objects.filter(name__icontains=q).first()
        if not match:
            raise Http404("Rezervație inexistentă")
        return match

@login_required
def home(request):
    stats = {
        "species": Species.objects.count(),
        "reserves": Reserve.objects.count(),
        "sites": Site.objects.count(),
        "habitats": Habitat.objects.count(),
    }

    context = {
        "stats": stats,
        "user_role": (
            "Administrator" if request.user.is_staff else
            ("Contributor" if request.user.groups.filter(name__iexact="Contributors").exists() else "Utilizator")
        ),
        "last_login": localtime(request.user.last_login) if request.user.last_login else None,
    }
    return render(request, "core/home.html", context)



@login_required
def coming_soon(request):
    return render(request, "core/coming_soon.html")



@login_required
def vizualizari_home(request):
    # HUB-ul cu cele 5 casete
    return render(request, "core/vizualizari_home.html")

def filtrari(request):
    """Public page with three stacked filter links, reusing vizualizări styles."""
    return render(request, "core/filtrari.html")

def filters_plante_rezervatii(request):
    """Wrapper that mirrors occurrences_filters_page behavior but renders a dedicated template under /filtrari/ namespace."""
    mode = (request.GET.get("mode") or "by_reserve_all").strip()
    reserve_name = (request.GET.get("reserve_name") or "").strip()
    raion = (request.GET.get("raion") or "").strip()

    all_reserves = Reserve.objects.order_by("name").only("id", "name")
    all_raions = (Reserve.objects
                  .exclude(raion__isnull=True).exclude(raion="")
                  .values_list("raion", flat=True).distinct().order_by("raion"))

    qs, error = _build_occurrence_filters_queryset(mode, reserve_name, raion)

    paginator, page_obj = _paginate(request, qs, default=50)
    rows = _rows_from_occurrences(page_obj.object_list)

    return render(request, "core/filters_plante_rezervatii.html", {
        "mode": mode,
        "reserve_name": reserve_name,
        "raion": raion,
        "all_reserves": all_reserves,
        "all_raions": all_raions,
        "rows": rows,
        "page_obj": page_obj,
        "paginator": paginator,
        "error": error,
    })

def _build_occurrence_filters_queryset(mode: str, reserve_name: str, raion: str):
    qs = (Occurrence.objects.select_related("species", "reserve").all())
    error = None
    if mode in ("by_reserve_all", "by_reserve_rare"):
        if not reserve_name:
            error = "Alege o rezervație."
            qs = qs.none()
        else:
            qs = qs.filter(reserve__name__iexact=reserve_name)
            if mode == "by_reserve_rare":
                qs = qs.filter(Q(is_rare=True) | Q(species__is_rare=True))
    elif mode in ("by_raion_all", "by_raion_rare"):
        if not raion:
            error = "Alege un raion."
            qs = qs.none()
        else:
            qs = qs.filter(reserve__raion__iexact=raion)
            if mode == "by_raion_rare":
                qs = qs.filter(Q(is_rare=True) | Q(species__is_rare=True))
    else:
        error = "Mod invalid."
        qs = qs.none()
    return qs.order_by("reserve__name", "species__denumire_stiintifica", "-year"), error

def _rows_from_occurrences(iterable):
    rows = []
    for o in iterable:
        rows.append({
            "reserve": o.reserve.name,
            "raion": o.reserve.raion or "",
            "species_sci": o.species.denumire_stiintifica,
            "species_pop": o.species.denumire_populara or "",
            "year": o.year,
            "rare": "Da" if (o.is_rare or o.species.is_rare) else "Nu",
            "lat": o.latitude,
            "lon": o.longitude,
        })
    return rows

def export_plante_rezervatii(request):
    kind = (request.GET.get("format") or request.GET.get("export") or "csv").lower()
    mode = (request.GET.get("mode") or "by_reserve_all").strip()
    reserve_name = (request.GET.get("reserve_name") or "").strip()
    raion = (request.GET.get("raion") or "").strip()

    qs, error = _build_occurrence_filters_queryset(mode, reserve_name, raion)
    if error:
        return HttpResponse(error, content_type="text/plain; charset=utf-8", status=400)

    headers = [
        "Rezervație", "Raion", "Specie (științific)", "Specie (popular)", "An", "Rară?", "Lat", "Lon"
    ]

    if kind == "xlsx":
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            kind = "csv"
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "PlanteRezervatii"
            ws.append(headers)
            for o in qs.iterator():
                ws.append([
                    o.reserve.name,
                    o.reserve.raion or "",
                    o.species.denumire_stiintifica,
                    o.species.denumire_populara or "",
                    o.year,
                    "Da" if (o.is_rare or o.species.is_rare) else "Nu",
                    o.latitude if o.latitude is not None else "",
                    o.longitude if o.longitude is not None else "",
                ])
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="plante_rezervatii.xlsx"'
            return resp

    # CSV fallback or explicit csv
    def rows_iter():
        yield ",".join(headers)
        import csv as _csv
        pseudo = _Echo()
        writer = _csv.writer(pseudo)
        for o in qs.iterator():
            yield writer.writerow([
                o.reserve.name,
                o.reserve.raion or "",
                o.species.denumire_stiintifica,
                o.species.denumire_populara or "",
                o.year,
                "Da" if (o.is_rare or o.species.is_rare) else "Nu",
                o.latitude if o.latitude is not None else "",
                o.longitude if o.longitude is not None else "",
            ])
    return StreamingHttpResponse(rows_iter(), content_type="text/csv; charset=utf-8", headers={
        "Content-Disposition": 'attachment; filename="plante_rezervatii.csv"'
    })

def filters_asociatii(request):
    mode = (request.GET.get("mode") or "by_reserve_year").strip()
    reserve_name = (request.GET.get("reserve_name") or "").strip()
    year_q = (request.GET.get("year") or "").strip()

    all_reserves = Reserve.objects.order_by("name").only("id", "name")

    links = ReserveAssociationYear.objects.select_related("reserve", "association")
    err = None

    if mode == "by_reserve_year":
        if not reserve_name or not year_q.isdigit():
            err = "Alege o rezervație și un an."
            qs = links.none()
        else:
            qs = links.filter(reserve__name__iexact=reserve_name, year=int(year_q)).order_by("association__name")
    elif mode == "by_reserve_all_years":
        if not reserve_name:
            err = "Alege o rezervație."
            qs = links.none()
        else:
            qs = links.filter(reserve__name__iexact=reserve_name).order_by("-year", "association__name")
    elif mode == "by_year_all_reserves":
        if not year_q.isdigit():
            err = "Alege un an."
            qs = links.none()
        else:
            qs = links.filter(year=int(year_q)).order_by("reserve__name", "association__name")
    else:
        err = "Mod invalid."
        qs = links.none()

    paginator, page_obj = _paginate(request, qs, default=50)
    rows = []
    for l in page_obj.object_list:
        rows.append({
            "reserve": l.reserve.name,
            "association": l.association.name,
            "year": l.year,
            "notes": l.notes or "",
        })

    return render(request, "core/filters_asociatii.html", {
        "mode": mode,
        "reserve_name": reserve_name,
        "year_q": year_q,
        "all_reserves": all_reserves,
        "rows": rows,
        "page_obj": page_obj,
        "paginator": paginator,
        "error": err,
    })

def export_asociatii(request):
    kind = (request.GET.get("format") or request.GET.get("export") or "csv").lower()
    mode = (request.GET.get("mode") or "by_reserve_year").strip()
    reserve_name = (request.GET.get("reserve_name") or "").strip()
    year_q = (request.GET.get("year") or "").strip()

    links = ReserveAssociationYear.objects.select_related("reserve", "association")

    if mode == "by_reserve_year":
        if not reserve_name or not year_q.isdigit():
            return HttpResponse("Alege o rezervație și un an.", content_type="text/plain; charset=utf-8", status=400)
        qs = links.filter(reserve__name__iexact=reserve_name, year=int(year_q)).order_by("association__name")
    elif mode == "by_reserve_all_years":
        if not reserve_name:
            return HttpResponse("Alege o rezervație.", content_type="text/plain; charset=utf-8", status=400)
        qs = links.filter(reserve__name__iexact=reserve_name).order_by("-year", "association__name")
    elif mode == "by_year_all_reserves":
        if not year_q.isdigit():
            return HttpResponse("Alege un an.", content_type="text/plain; charset=utf-8", status=400)
        qs = links.filter(year=int(year_q)).order_by("reserve__name", "association__name")
    else:
        return HttpResponse("Mod invalid.", content_type="text/plain; charset=utf-8", status=400)

    headers = ["Rezervație", "Asociație", "An", "Note"]

    if kind == "xlsx":
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            kind = "csv"
        else:
            wb = Workbook(); ws = wb.active; ws.title = "Asociatii"; ws.append(headers)
            for l in qs.iterator():
                ws.append([l.reserve.name, l.association.name, l.year, l.notes or ""])
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="asociatii.xlsx"'
            return resp

    def rows_iter():
        import csv as _csv
        pseudo = _Echo(); writer = _csv.writer(pseudo)
        yield writer.writerow(headers)
        for l in qs.iterator():
            yield writer.writerow([l.reserve.name, l.association.name, l.year, l.notes or ""])
    return StreamingHttpResponse(rows_iter(), content_type="text/csv; charset=utf-8", headers={
        "Content-Disposition": 'attachment; filename="asociatii.csv"'
    })

def filters_situri_habitat(request):
    """Wrapper that mirrors sitehab_filters_page behavior but renders under /filtrari/ namespace."""
    mode = request.GET.get("mode", "by_site")
    site_name = (request.GET.get("site_name") or "").strip()
    habitat_name = (request.GET.get("habitat_name") or "").strip()
    year = (request.GET.get("year") or "").strip()
    export = (request.GET.get("export") or "").strip().lower()

    qs_full = SiteHabitat.objects.select_related("site", "habitat")

    if mode == "by_site" and site_name:
        qs_full = qs_full.filter(site__name__iexact=site_name).order_by(
            "year", "habitat__name_romanian", "habitat__name_english"
        )
        title = f"Habitate în site-ul: {site_name}"
    elif mode == "by_habitat" and habitat_name:
        qs_full = qs_full.filter(
            Q(habitat__name_romanian__iexact=habitat_name) |
            Q(habitat__name_english__iexact=habitat_name) |
            Q(habitat__code__iexact=habitat_name)
        ).order_by("year", "site__name")
        title = f"Site-uri pentru habitat: {habitat_name}"
    elif mode == "by_year" and year.isdigit():
        qs_full = qs_full.filter(year=int(year)).order_by(
            "site__name", "habitat__name_romanian", "habitat__name_english"
        )
        title = f"Relații Site–Habitat în anul: {year}"
    else:
        qs_full = qs_full.none()
        title = "Selectează un filtru"

    if export in ("csv", "xlsx"):
        return _export_sitehab(qs_full, export)

    paginator, page_obj = _paginate(request, qs_full, default=50)
    qs = page_obj.object_list

    sites = Site.objects.order_by("name").values_list("name", flat=True)
    habitats = Habitat.objects.order_by("name_romanian", "name_english").values_list("name_romanian", flat=True)

    return render(request, "core/filters_situri_habitat.html", {
        "mode": mode,
        "site_name": site_name,
        "habitat_name": habitat_name,
        "year": year,
        "qs": qs,
        "page_obj": page_obj,
        "paginator": paginator,
        "title": title,
        "sites": list(sites),
        "habitats": list(habitats),
    })

@login_required
def coming_soon(request):
    return render(request, "core/coming_soon.html")

@login_required
def viz_specii(request):
    """Listă cu căutare tolerantă (case/diacritice) și paginare pe carduri."""
    q = (request.GET.get("q") or "").strip()
    qs = Species.objects.all()

    def _normalize_text(value: str) -> str:
        if not value:
            return ""
        decomposed = unicodedata.normalize("NFKD", str(value))
        no_accents = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
        return no_accents.lower().strip()

    def _fuzzy_ratio(a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        return difflib.SequenceMatcher(None, a, b).ratio()

    if not q:
        qs = qs.order_by("denumire_stiintifica")
        paginator, page_obj = _paginate(request, qs, default=24)
    else:
        norm_q = _normalize_text(q)
        candidates = list(qs.only(
            "id", "denumire_stiintifica", "denumire_populara", "familia", "clasa", "habitat", "localitatea"
        ))
        scored = []
        for s in candidates:
            fields = [
                _normalize_text(s.denumire_stiintifica),
                _normalize_text(s.denumire_populara),
                _normalize_text(s.familia),
                _normalize_text(s.clasa),
                _normalize_text(s.habitat),
                _normalize_text(s.localitatea),
            ]
            best = 0.0
            substr_bonus = 0.0
            for f in fields:
                if not f:
                    continue
                if norm_q in f:
                    substr_bonus = max(substr_bonus, 0.15)
                best = max(best, _fuzzy_ratio(norm_q, f))
            total_score = best + substr_bonus
            if total_score >= 0.55:
                scored.append((total_score, s))

        scored.sort(key=lambda t: (-t[0], _normalize_text(t[1].denumire_stiintifica)))
        ordered = [s for _, s in scored]
        paginator, page_obj = _paginate(request, ordered, default=24)

    return render(request, "core/viz_specii.html", {
        "page_obj": page_obj,
        "paginator": paginator,
        "q": q,
    })

@login_required
def viz_specii_detail(request, pk: int):
    sp = get_object_or_404(Species, pk=pk)
    points_qs = (Occurrence.objects
                 .filter(species=sp)
                 .exclude(latitude__isnull=True)
                 .exclude(longitude__isnull=True))

    # transformăm într-o listă simplă pentru template (lat, lon, label)
    points = []
    for o in points_qs:
        try:
            lat = float(o.latitude)
            lon = float(o.longitude)
        except (TypeError, ValueError):
            continue
        label = f"{sp.denumire_stiintifica} — {o.reserve.name} ({o.year})" if getattr(o, "reserve", None) else sp.denumire_stiintifica
        points.append({
            "lat": lat,
            "lon": lon,
            "label": label,
        })

    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    return render(request, "core/viz_specii_detail.html", {
        "sp": sp,
        "points": points,
        "is_admin": is_admin,
    })

@login_required
def viz_rezervatii(request):
    """Listă de rezervații în carduri, cu căutare tolerantă (diacritice/typo)."""
    q = (request.GET.get("q") or "").strip()
    qs = Reserve.objects.all()

    def _normalize_text(value: str) -> str:
        if not value:
            return ""
        decomposed = unicodedata.normalize("NFKD", str(value))
        no_accents = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
        return no_accents.lower().strip()

    def _fuzzy_ratio(a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        return difflib.SequenceMatcher(None, a, b).ratio()

    if not q:
        qs = qs.order_by("name")
        paginator, page_obj = _paginate(request, qs, default=24)
    else:
        norm_q = _normalize_text(q)
        candidates = list(qs.only(
            "id", "name", "raion", "amplasare", "proprietar", "category", "subcategory"
        ))
        scored = []
        for r in candidates:
            fields = [
                _normalize_text(r.name),
                _normalize_text(r.raion),
                _normalize_text(r.amplasare),
                _normalize_text(r.proprietar),
                _normalize_text(r.category),
                _normalize_text(r.subcategory),
            ]
            best = 0.0
            substr_bonus = 0.0
            for f in fields:
                if not f:
                    continue
                if norm_q in f:
                    substr_bonus = max(substr_bonus, 0.15)
                best = max(best, _fuzzy_ratio(norm_q, f))
            total_score = best + substr_bonus
            if total_score >= 0.55:
                scored.append((total_score, r))

        scored.sort(key=lambda t: (-t[0], _normalize_text(t[1].name)))
        ordered = [r for _, r in scored]
        paginator, page_obj = _paginate(request, ordered, default=24)

    return render(request, "core/viz_rezervatii.html", {
        "page_obj": page_obj,
        "paginator": paginator,
        "q": q,
    })

@login_required
def viz_asociatii(request):
    """Listă de asociații cu căutare tolerantă (diacritice/typo) și paginare, ca la Rezervații.

    Implementare fuzzy în Postgres: unaccent + trigram similarity/word-similarity.
    """
    q = (request.GET.get("q") or "").strip()
    qs = Association.objects.all()

    if not q:
        qs = qs.order_by("name")
        paginator, page_obj = _paginate(request, qs, default=24)
    else:
        # Normalize query (lowercase). Unaccent is applied on both sides at SQL level
        q_norm = (q or "").lower().strip()

        # Use PostgreSQL unaccent
        class Unaccent(Func):
            function = "unaccent"
            output_field = CharField()

        name_u = Unaccent(Lower(F("name")))
        q_u = Unaccent(Lower(Value(q_norm, output_field=CharField())))

        qs = qs.annotate(
            name_u=name_u,
            sim=TrigramSimilarity(name_u, q_u),
            starts=Case(
                When(name_u__startswith=q_norm, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
        ).filter(
            Q(name_u__icontains=q_norm) | Q(sim__gt=0.16)
        ).order_by("-starts", "-sim", "name")

        paginator, page_obj = _paginate(request, qs, default=24)

    return render(request, "core/viz_asociatii.html", {
        "page_obj": page_obj,
        "paginator": paginator,
        "q": q,
    })

@login_required
def viz_rezervatii_detail(request, pk: int):
    r = get_object_or_404(Reserve, pk=pk)
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    return render(request, "core/viz_rezervatii_detail.html", {"r": r, "is_admin": is_admin})

@login_required
def viz_asociatii_detail(request, pk: int):
    a = get_object_or_404(Association, pk=pk)
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    return render(request, "core/viz_asociatii_detail.html", {"a": a, "is_admin": is_admin})

@login_required
def update_species_description(request, pk: int):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Method not allowed"}, status=405)
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    if not is_admin:
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)
    species = get_object_or_404(Species, pk=pk)
    raw = request.POST.get("description", "")
    safe = strip_tags(raw)
    species.description = safe
    species.save(update_fields=["description", "updated_at"])
    return JsonResponse({"ok": True, "description": safe})

@login_required
def update_species_meta(request, pk: int):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Method not allowed"}, status=405)
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    if not is_admin:
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)
    
    species = get_object_or_404(Species, pk=pk)
    
    # Fields that can be updated
    updatable_fields = [
        'denumire_stiintifica', 'denumire_populara', 'clasa', 'familia', 'habitat', 'localitatea',
        'silvice', 'pajisti_sau_stepice', 'stancarii', 'palustre_si_acvatice',
        'conventia_berna', 'directiva_habitate', 'cartea_rosie', 'cartea_rosie_cat',
        'frecventa', 'notes'
    ]
    
    changed = {}
    update_fields = ['updated_at']
    
    for field in updatable_fields:
        if field in request.POST:
            value = request.POST.get(field, '').strip()
            
            # Handle boolean fields
            if field in ['silvice', 'pajisti_sau_stepice', 'stancarii', 'palustre_si_acvatice', 'conventia_berna', 'directiva_habitate']:
                new_value = value.lower() in ['true', '1', 'yes', 'on']
                if getattr(species, field) != new_value:
                    setattr(species, field, new_value)
                    changed[field] = new_value
                    update_fields.append(field)
            
            # Handle integer fields
            elif field == 'cartea_rosie':
                try:
                    new_value = int(value) if value else None
                    if getattr(species, field) != new_value:
                        setattr(species, field, new_value)
                        changed[field] = new_value
                        update_fields.append(field)
                except ValueError:
                    pass  # Skip invalid integer values
            
            # Handle scientific name with validation
            elif field == 'denumire_stiintifica':
                new_raw = value
                # Trim and collapse spaces
                new_norm = ' '.join(new_raw.split()) if new_raw else None
                if new_norm:
                    # Capitalize genus (first token)
                    parts = new_norm.split(' ')
                    if parts:
                        parts[0] = parts[0][:1].upper() + parts[0][1:].lower()
                    new_norm = ' '.join(parts)
                    # Length validation 3–200
                    if not (3 <= len(new_norm) <= 200):
                        return JsonResponse({"ok": False, "error": "Scientific name length must be 3–200"}, status=400)
                # Uniqueness check (case-insensitive)
                if new_norm and Species.objects.filter(denumire_stiintifica__iexact=new_norm).exclude(pk=species.pk).exists():
                    return JsonResponse({"ok": False, "error": "Scientific name already exists"}, status=400)
                if getattr(species, field) != new_norm:
                    setattr(species, field, new_norm)
                    changed[field] = new_norm
                    update_fields.append(field)

            # Handle other text fields
            else:
                new_value = value if value else None
                if getattr(species, field) != new_value:
                    setattr(species, field, new_value)
                    changed[field] = new_value
                    update_fields.append(field)
    
    if changed:
        species.save(update_fields=update_fields)
    
    return JsonResponse({"ok": True, "changed": changed})

@login_required
@require_POST
def update_association_meta(request, pk: int):
    """Update Association fields (admins only)."""
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    if not is_admin:
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)

    a = get_object_or_404(Association, pk=pk)

    def norm_text(v):
        return strip_tags((v or "").strip()) or None

    changed = {}
    # Current Association model has only 'name' and 'notes'
    for field in ("name", "notes"):
        if field in request.POST:
            val = norm_text(request.POST.get(field))
            if getattr(a, field) != val:
                setattr(a, field, val)
                changed[field] = val

    if not changed:
        return JsonResponse({"ok": True, "changed": {}}, status=200)

    a.save()
    return JsonResponse({"ok": True, "changed": changed})

@login_required
def update_reserve_description(request, pk: int):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Method not allowed"}, status=405)
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    if not is_admin:
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)
    reserve = get_object_or_404(Reserve, pk=pk)
    raw = request.POST.get("description", "")
    safe = strip_tags(raw)
    reserve.description = safe
    reserve.save(update_fields=["description", "updated_at"])
    return JsonResponse({"ok": True, "description": safe})


@login_required
@require_POST
def update_reserve_meta(request, pk: int):
    """Update selected Reserve fields. Admins only. Returns changed fields."""
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    if not is_admin:
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)

    reserve = get_object_or_404(Reserve, pk=pk)

    def norm_text(v):
        return strip_tags((v or "").strip())

    changed = {}
    # Text fields
    for field in ("name", "raion", "amplasare", "proprietar", "subcategory", "category"):
        if field in request.POST:
            val = norm_text(request.POST.get(field))
            if getattr(reserve, field) != val:
                setattr(reserve, field, val or None)
                changed[field] = val

    # Numeric area
    if "suprafata_ha" in request.POST:
        raw = (request.POST.get("suprafata_ha") or "").strip()
        if raw != "":
            try:
                num = float(raw)
                if num < 0:
                    return JsonResponse({"ok": False, "error": "Area must be positive"}, status=400)
            except ValueError:
                return JsonResponse({"ok": False, "error": "Invalid area"}, status=400)
        else:
            num = None
        if reserve.suprafata_ha != num:
            reserve.suprafata_ha = num
            changed["suprafata_ha"] = num

    # Coordinates (accept both 'latitude/longitude' and 'latitudine/longitudine')
    lat = request.POST.get("latitude")
    lon = request.POST.get("longitude")
    if lat is None and "latitudine" in request.POST:
        lat = request.POST.get("latitudine")
    if lon is None and "longitudine" in request.POST:
        lon = request.POST.get("longitudine")
    if lat is not None or lon is not None:
        def parse_or_none(x):
            x = (x or "").strip()
            return None if x == "" else float(x)
        try:
            lat_v = parse_or_none(lat)
            lon_v = parse_or_none(lon)
            if lat_v is not None and not (-90.0 <= lat_v <= 90.0):
                return JsonResponse({"ok": False, "error": "Latitude out of range"}, status=400)
            if lon_v is not None and not (-180.0 <= lon_v <= 180.0):
                return JsonResponse({"ok": False, "error": "Longitude out of range"}, status=400)
        except ValueError:
            return JsonResponse({"ok": False, "error": "Invalid coordinates"}, status=400)
        if reserve.latitude != lat_v:
            reserve.latitude = lat_v
            changed["latitude"] = lat_v
        if reserve.longitude != lon_v:
            reserve.longitude = lon_v
            changed["longitude"] = lon_v

    if not changed:
        return JsonResponse({"ok": True, "changed": {}}, status=200)

    reserve.save()
    return JsonResponse({"ok": True, "changed": changed})

@login_required
def viz_situri(request):
    """Listă de site-uri, cu căutare tolerantă ca la Rezervații."""
    q = (request.GET.get("q") or "").strip()
    qs = Site.objects.all()

    if not q:
        qs = qs.order_by("name")
        paginator, page_obj = _paginate(request, qs, default=24)
    else:
        q_norm = (q or "").lower().strip()

        class Unaccent(Func):
            function = "unaccent"
            output_field = CharField()

        def unaccent_lower(expr):
            return Lower(Unaccent(expr))

        name_u = unaccent_lower(F("name"))
        # Adapt fields as needed when available
        qs = qs.annotate(
            name_u=name_u,
            sim=TrigramSimilarity(name_u, Unaccent(Lower(Value(q_norm, output_field=CharField())))),
            starts=Case(
                When(name_u__startswith=q_norm, then=Value(1)),
                default=Value(0), output_field=IntegerField(),
            ),
        ).filter(
            Q(name_u__icontains=q_norm) | Q(sim__gt=0.16)
        ).order_by("-starts", "-sim", "name")

        paginator, page_obj = _paginate(request, qs, default=24)

    return render(request, "core/viz_situri_list.html", {
        "page_obj": page_obj,
        "paginator": paginator,
        "q": q,
    })

@login_required
def viz_habitate(request):
    """Listă de habitate cu căutare tolerantă (diacritice/typo) și paginare, ca la Rezervații.

    Implementare fuzzy în Postgres: unaccent + trigram similarity/word-similarity.
    """
    q = (request.GET.get("q") or "").strip()
    qs = Habitat.objects.all()

    if not q:
        qs = qs.order_by("name_romanian")
        paginator, page_obj = _paginate(request, qs, default=24)
    else:
        # Normalize query (lowercase). Unaccent is applied on both sides at SQL level
        q_norm = (q or "").lower().strip()

        # Use PostgreSQL unaccent
        class Unaccent(Func):
            function = "unaccent"
            output_field = CharField()

        name_ro_u = Unaccent(Lower(F("name_romanian")))
        name_en_u = Unaccent(Lower(F("name_english")))
        code_u = Unaccent(Lower(F("code")))
        q_u = Unaccent(Lower(Value(q_norm, output_field=CharField())))

        qs = qs.annotate(
            name_ro_u=name_ro_u,
            name_en_u=name_en_u,
            code_u=code_u,
            sim_ro=TrigramSimilarity(name_ro_u, q_u),
            sim_en=TrigramSimilarity(name_en_u, q_u),
            sim_code=TrigramSimilarity(code_u, q_u),
            wsim_ro=TrigramWordSimilarity(q_u, name_ro_u) if TrigramWordSimilarity else Value(0.0, output_field=FloatField()),
            wsim_en=TrigramWordSimilarity(q_u, name_en_u) if TrigramWordSimilarity else Value(0.0, output_field=FloatField()),
            starts_ro=Case(
                When(name_ro_u__startswith=q_norm, then=1),
                default=0,
                output_field=IntegerField(),
            ),
            starts_en=Case(
                When(name_en_u__startswith=q_norm, then=1),
                default=0,
                output_field=IntegerField(),
            ),
            starts_code=Case(
                When(code_u__startswith=q_norm, then=1),
                default=0,
                output_field=IntegerField(),
            ),
        ).filter(
            Q(name_ro_u__icontains=q_norm) | 
            Q(name_en_u__icontains=q_norm) | 
            Q(code_u__icontains=q_norm) |
            Q(sim_ro__gt=0.25) | 
            Q(sim_en__gt=0.25) | 
            Q(sim_code__gt=0.25) |
            Q(wsim_ro__gt=0.20) | 
            Q(wsim_en__gt=0.20)
        ).order_by(
            "-starts_ro", "-starts_en", "-starts_code",
            Greatest("wsim_ro", "wsim_en", "sim_ro", "sim_en", "sim_code").desc(),
            "name_romanian"
        )

        paginator, page_obj = _paginate(request, qs, default=24)

    return render(request, "core/viz_habitate.html", {
        "page_obj": page_obj,
        "paginator": paginator,
        "q": q,
    })

@login_required
def viz_habitate_detail(request, pk: int):
    h = get_object_or_404(Habitat, pk=pk)
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    return render(request, "core/viz_habitate_detail.html", {"h": h, "is_admin": is_admin})

@login_required
@require_POST
def update_habitat_meta(request, pk: int):
    """Update Habitat fields (admins only)."""
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    if not is_admin:
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)

    h = get_object_or_404(Habitat, pk=pk)

    def norm_text(v):
        return strip_tags((v or "").strip()) or None

    changed = {}
    # Habitat model fields: name_romanian, name_english, code, notes
    for field in ("name_romanian", "name_english", "code", "notes"):
        if field in request.POST:
            val = norm_text(request.POST.get(field))
            if getattr(h, field) != val:
                setattr(h, field, val)
                changed[field] = val

    if not changed:
        return JsonResponse({"ok": True, "changed": {}}, status=200)

    h.save()
    return JsonResponse({"ok": True, "changed": changed})

@login_required
def comparatii_home(request):
    # stub simplu până definim comparațiile
    return render(request, "core/coming_soon.html", {
        "title": "Comparații",
        "subtitle": "Compară rezervații, habitate sau ani diferiți."
    })

@login_required
def species_list(request):
    return HttpResponse("Test species_list – aici vor veni filtrările pentru specii.")


@login_required
@permission_required("core.view_occurrence", raise_exception=True)
def occurrences_filters_page(request):
    """
    4 moduri:
      - by_reserve_all:    plante (toate) din rezervația X
      - by_reserve_rare:   plante rare din rezervația X
      - by_raion_all:      plante (toate) pe raionul R
      - by_raion_rare:     plante rare pe raionul R
    GET:
      mode: vezi mai sus (default by_reserve_all)
      reserve_name: necesar în modurile 'by_reserve_*' (după nume, nu ID)
      raion:        necesar în modurile 'by_raion_*'
    """
    mode = (request.GET.get("mode") or "by_reserve_all").strip()
    reserve_name = (request.GET.get("reserve_name") or "").strip()
    raion = (request.GET.get("raion") or "").strip()

    # dropdown-uri
    all_reserves = Reserve.objects.order_by("name").only("id", "name")
    all_raions = (Reserve.objects
                  .exclude(raion__isnull=True).exclude(raion="")
                  .values_list("raion", flat=True).distinct().order_by("raion"))

    qs = (Occurrence.objects
          .select_related("species", "reserve")
          .all())

    error = None

    if mode in ("by_reserve_all", "by_reserve_rare"):
        if not reserve_name:
            error = "Alege o rezervație."
            qs = qs.none()
        else:
            qs = qs.filter(reserve__name__iexact=reserve_name)
            if mode == "by_reserve_rare":
                qs = qs.filter(Q(is_rare=True) | Q(species__is_rare=True))

    elif mode in ("by_raion_all", "by_raion_rare"):
        if not raion:
            error = "Alege un raion."
            qs = qs.none()
        else:
            qs = qs.filter(reserve__raion__iexact=raion)
            if mode == "by_raion_rare":
                qs = qs.filter(Q(is_rare=True) | Q(species__is_rare=True))

    else:
        error = "Mod invalid."
        qs = qs.none()

    # Sortare prietenoasă: după rezervație, apoi specie, apoi an desc
    qs = qs.order_by("reserve__name", "species__denumire_stiintifica", "-year")

    # Paginăm queryset-ul și construim rândurile DOAR pentru pagina curentă
    paginator, page_obj = _paginate(request, qs, default=50)
    rows = []
    for o in page_obj.object_list:
        rows.append({
            "reserve": o.reserve.name,
            "raion": o.reserve.raion or "",
            "species_sci": o.species.denumire_stiintifica,
            "species_pop": o.species.denumire_populara or "",
            "year": o.year,
            "rare": "Da" if (o.is_rare or o.species.is_rare) else "Nu",
            "lat": o.latitude,
            "lon": o.longitude,
        })

    return render(request, "core/occurrences_filters.html", {
        "mode": mode,
        "reserve_name": reserve_name,
        "raion": raion,
        "all_reserves": all_reserves,
        "all_raions": all_raions,
        "rows": rows,
        "page_obj": page_obj,
        "paginator": paginator,
        "error": error,
    })


@login_required
@permission_required("core.view_sitehabitat", raise_exception=True)
def sitehab_filters_page(request):
    """
    Mode:
      - by_site:     toate habitatele din site-ul X (param: site_name)
      - by_habitat:  toate site-urile care conțin habitatul H (param: habitat_name)
      - by_year:     toate relațiile din anul Y (param: year)
    Export:
      - ?export=csv  sau  ?export=xlsx
    """
    mode = request.GET.get("mode", "by_site")
    site_name = (request.GET.get("site_name") or "").strip()
    habitat_name = (request.GET.get("habitat_name") or "").strip()
    year = (request.GET.get("year") or "").strip()
    export = (request.GET.get("export") or "").strip().lower()

    qs_full = SiteHabitat.objects.select_related("site", "habitat")

    if mode == "by_site" and site_name:
        qs_full = qs_full.filter(site__name__iexact=site_name).order_by(
            "year", "habitat__name_romanian", "habitat__name_english"
        )
        title = f"Habitate în site-ul: {site_name}"
    elif mode == "by_habitat" and habitat_name:
        qs_full = qs_full.filter(
            Q(habitat__name_romanian__iexact=habitat_name) |
            Q(habitat__name_english__iexact=habitat_name) |
            Q(habitat__code__iexact=habitat_name)
        ).order_by("year", "site__name")
        title = f"Site-uri pentru habitat: {habitat_name}"
    elif mode == "by_year" and year.isdigit():
        qs_full = qs_full.filter(year=int(year)).order_by(
            "site__name", "habitat__name_romanian", "habitat__name_english"
        )
        title = f"Relații Site–Habitat în anul: {year}"
    else:
        qs_full = qs_full.none()
        title = "Selectează un filtru"

    # Export (din QS-ul complet, nu doar pagina curentă)
    if export in ("csv", "xlsx"):
        return _export_sitehab(qs_full, export)

    # Paginăm pentru afișare
    paginator, page_obj = _paginate(request, qs_full, default=50)
    qs = page_obj.object_list  # pentru compatibilitate cu template-ul existent

    sites = Site.objects.order_by("name").values_list("name", flat=True)
    habitats = Habitat.objects.order_by("name_romanian", "name_english").values_list("name_romanian", flat=True)

    return render(request, "core/sitehab_filters.html", {
        "mode": mode,
        "site_name": site_name,
        "habitat_name": habitat_name,
        "year": year,
        "qs": qs,
        "page_obj": page_obj,
        "paginator": paginator,
        "title": title,
        "sites": list(sites),
        "habitats": list(habitats),
    })
    """
    Mode:
      - by_site:     toate habitatele din site-ul X (param: site_name)
      - by_habitat:  toate site-urile care conțin habitatul H (param: habitat_name)
      - by_year:     toate relațiile din anul Y (param: year)
    Export:
      - ?export=csv  sau  ?export=xlsx
    """
    mode = request.GET.get("mode", "by_site")
    site_name = (request.GET.get("site_name") or "").strip()
    habitat_name = (request.GET.get("habitat_name") or "").strip()
    year = (request.GET.get("year") or "").strip()
    export = (request.GET.get("export") or "").strip().lower()

    qs = SiteHabitat.objects.select_related("site", "habitat")

    if mode == "by_site" and site_name:
        qs = qs.filter(site__name__iexact=site_name).order_by("year", "habitat__name_romanian", "habitat__name_english")
        title = f"Habitate în site-ul: {site_name}"
    elif mode == "by_habitat" and habitat_name:
        qs = qs.filter(
            Q(habitat__name_romanian__iexact=habitat_name) |
            Q(habitat__name_english__iexact=habitat_name) |
            Q(habitat__code__iexact=habitat_name)
        ).order_by("year", "site__name")
        title = f"Site-uri pentru habitat: {habitat_name}"
    elif mode == "by_year" and year.isdigit():
        qs = qs.filter(year=int(year)).order_by("site__name", "habitat__name_romanian", "habitat__name_english")
        title = f"Relații Site–Habitat în anul: {year}"
    else:
        qs = qs.none()
        title = "Selectează un filtru"

    if export in ("csv", "xlsx"):
        return _export_sitehab(qs, export)

    sites = Site.objects.order_by("name").values_list("name", flat=True)
    habitats = Habitat.objects.order_by("name_romanian", "name_english").values_list("name_romanian", flat=True)

    return render(request, "core/sitehab_filters.html", {
        "mode": mode,
        "site_name": site_name,
        "habitat_name": habitat_name,
        "year": year,
        "qs": qs,
        "title": title,
        "sites": list(sites),
        "habitats": list(habitats),
    })

@login_required
def comparatii_home(request):
    return render(request, "core/comparatii_home.html")


@login_required
def comparatii_plante_list(request):
    """List similar to vizualizări/rezervatii but for starting rare plant comparisons.

    Reuses the same dataset and pagination pattern as viz_rezervatii.
    """
    q = (request.GET.get("q") or "").strip()
    qs = Reserve.objects.all()

    if not q:
        qs = qs.order_by("name")
        paginator, page_obj = _paginate(request, qs, default=24)
    else:
        # Mirror basic tolerant search from viz_rezervatii for parity
        def _normalize_text(value: str) -> str:
            if not value:
                return ""
            decomposed = unicodedata.normalize("NFKD", str(value))
            no_accents = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
            return no_accents.lower().strip()

        def _fuzzy_ratio(a: str, b: str) -> float:
            if not a or not b:
                return 0.0
            return difflib.SequenceMatcher(None, a, b).ratio()

        norm_q = _normalize_text(q)
        candidates = list(qs.only(
            "id", "name", "raion", "amplasare", "proprietar", "category", "subcategory"
        ))
        scored = []
        for r in candidates:
            fields = [
                _normalize_text(r.name),
                _normalize_text(r.raion),
                _normalize_text(r.amplasare),
                _normalize_text(r.proprietar),
                _normalize_text(r.category),
                _normalize_text(r.subcategory),
            ]
            best = 0.0
            substr_bonus = 0.0
            for f in fields:
                if not f:
                    continue
                if norm_q in f:
                    substr_bonus = max(substr_bonus, 0.15)
                best = max(best, _fuzzy_ratio(norm_q, f))
            total_score = best + substr_bonus
            if total_score >= 0.55:
                scored.append((total_score, r))

        scored.sort(key=lambda t: (-t[0], _normalize_text(t[1].name)))
        ordered = [r for _, r in scored]
        paginator, page_obj = _paginate(request, ordered, default=24)

    all_reserves = Reserve.objects.order_by("name").only("id", "name", "raion")
    return render(request, "core/comparatii_plante.html", {
        "page_obj": page_obj,
        "paginator": paginator,
        "q": q,
        "all_reserves": all_reserves,
    })


@login_required
def comparatii_plante_detail(request, pk: int):
    r = get_object_or_404(Reserve, pk=pk)
    mode = (request.GET.get("mode") or "years").strip()
    
    # Parse base reserve and year from URL
    base_reserve_id = request.GET.get("base_reserve")
    base_year = request.GET.get("base_year")
    if base_reserve_id and base_reserve_id.isdigit():
        base_reserve_id = int(base_reserve_id)
    else:
        base_reserve_id = None
    if base_year and base_year.isdigit():
        base_year = int(base_year)
    else:
        base_year = None
    # Available years for rare species occurrences in this reserve
    years_qs = (Occurrence.objects
                .filter(reserve_id=pk)
                .filter(Q(is_rare=True) | Q(species__is_rare=True))
                .exclude(year__isnull=True)
                .values_list("year", flat=True)
                .distinct())
    available_years = sorted(set(int(y) for y in years_qs if y), reverse=True)

    # Parse selected years from URL, keep only available, limit 2..4
    years_param = (request.GET.get("years") or "").strip()
    selected_years = []
    if years_param:
        for part in years_param.split(','):
            part = part.strip()
            if part.isdigit():
                y = int(part)
                if y in available_years:
                    selected_years.append(y)
    # keep stable order as provided, but also ensure uniqueness
    seen = set()
    selected_years = [y for y in selected_years if not (y in seen or seen.add(y))][:4]

    has_valid_selection = 2 <= len(selected_years) <= 4 if mode == "years" else False

    rows = []
    page_obj = None
    paginator = None
    total_species = 0

    if mode == "years" and has_valid_selection:
        # Fetch occurrences once for selected years
        occ = (Occurrence.objects
               .filter(reserve_id=pk, year__in=selected_years)
               .filter(Q(is_rare=True) | Q(species__is_rare=True))
               .values_list("species_id", "year")
               .distinct())

        # Build presence map: species_id -> set(years)
        species_to_years = {}
        for sid, yr in occ:
            if sid is None or yr is None:
                continue
            species_to_years.setdefault(int(sid), set()).add(int(yr))

        total_species = len(species_to_years)

        # Order species by scientific name
        species_ids = list(species_to_years.keys())
        sp_qs = Species.objects.filter(id__in=species_ids).only("id", "denumire_stiintifica").order_by("denumire_stiintifica")
        ordered_species = list(sp_qs)

        # Build rows as dicts for template
        for sp in ordered_species:
            presence = []
            years_set = species_to_years.get(sp.id, set())
            for y in selected_years:
                presence.append(y in years_set)
            rows.append({
                "species_id": sp.id,
                "species_name": sp.denumire_stiintifica,
                "presence": presence,
            })

        # Paginate rows if large
        paginator, page_obj = _paginate(request, rows, default=50)
        rows = page_obj.object_list

    # Mode: reserves (compare with other reserves)
    columns = []
    rows_res = []
    if mode == "reserves":
        # Expect base=pk:year and res=rid:year,...
        base_pair = (request.GET.get("base") or "").strip()
        base_year = None
        if ":" in base_pair:
            b_id, b_y = base_pair.split(":", 1)
            if b_id.strip().isdigit() and int(b_id.strip()) == pk and b_y.strip().isdigit():
                base_year = int(b_y.strip())

        res_param = (request.GET.get("res") or "").strip()
        other_pairs = []
        if res_param:
            for part in res_param.split(','):
                if ":" in part:
                    rid_s, y_s = part.split(":", 1)
                    rid_s = rid_s.strip(); y_s = y_s.strip()
                    if rid_s.isdigit() and y_s.isdigit():
                        rid = int(rid_s); yy = int(y_s)
                        if rid != pk:
                            other_pairs.append((rid, yy))

        valid_pairs = []
        if base_year is not None:
            valid_pairs.append((pk, base_year))
        for rid, yy in other_pairs[:3]:
            valid_pairs.append((rid, yy))

        # Validate years availability per reserve (rare-only)
        ids_all = list({rid for rid, _ in valid_pairs})
        years_by_rid = {}
        if ids_all:
            for rid in ids_all:
                ys = (Occurrence.objects
                      .filter(reserve_id=rid)
                      .filter(Q(is_rare=True) | Q(species__is_rare=True))
                      .exclude(year__isnull=True)
                      .values_list("year", flat=True)
                      .distinct())
                years_by_rid[rid] = set(int(y) for y in ys if y)

            valid_pairs = [(rid, yy) for (rid, yy) in valid_pairs if yy in years_by_rid.get(rid, set())]

        if len(valid_pairs) >= 2:
            rid_list = [rid for rid, _ in valid_pairs]
            years_list = [yy for _, yy in valid_pairs]
            occ = (Occurrence.objects
                   .filter(reserve_id__in=rid_list, year__in=years_list)
                   .filter(Q(is_rare=True) | Q(species__is_rare=True))
                   .values_list("species_id", "reserve_id", "year")
                   .distinct())
            presence_map = {}
            for sid, rid, yy in occ:
                if sid is None or rid is None or yy is None:
                    continue
                presence_map.setdefault(int(sid), set()).add((int(rid), int(yy)))

            species_ids = list(presence_map.keys())
            sp_qs = Species.objects.filter(id__in=species_ids).only("id", "denumire_stiintifica").order_by("denumire_stiintifica")
            ordered_species = list(sp_qs)

            rid_to_name = { rr.id: rr.name for rr in Reserve.objects.filter(id__in=set(rid_list)).only("id","name") }

            # Ensure base first, then others in provided order
            pairs_ordered = []
            for rid, yy in valid_pairs:
                if rid == pk:
                    pairs_ordered.append((rid, yy))
                    break
            for rid, yy in valid_pairs:
                if rid != pk:
                    pairs_ordered.append((rid, yy))

            for rid, yy in pairs_ordered:
                columns.append({"reserve_id": rid, "reserve_name": rid_to_name.get(rid, f"#{rid}"), "year": yy})

            for sp in ordered_species:
                presence = []
                pset = presence_map.get(sp.id, set())
                for col in columns:
                    presence.append((col["reserve_id"], col["year"]) in pset)
                rows_res.append({
                    "species_id": sp.id,
                    "species_name": sp.denumire_stiintifica,
                    "presence": presence,
                })

    # Reserves combobox data (exclude current reserve)
    all_reserves = Reserve.objects.exclude(id=pk).order_by("name").only("id", "name")

    context = {
        "r": r,
        "available_years": available_years,
        "selected_years": selected_years,
        "has_valid_selection": has_valid_selection,
        "mode": mode,
        "rows": rows,
        "page_obj": page_obj,
        "paginator": paginator,
        "total_species": total_species,
        # reserves-mode context
        "columns": columns,
        "rows_res": rows_res,
        "all_reserves": all_reserves,
        "base_reserve_id_url": base_reserve_id,
        "base_year_url": base_year,
    }
    return render(request, "core/comparatii_plante_detail.html", context)


def _species_ids_for_reserve(reserve_id: int, only_rare: bool = True):
    qs = (Occurrence.objects
          .filter(reserve_id=reserve_id)
          .select_related("species"))
    if only_rare:
        qs = qs.filter(Q(is_rare=True) | Q(species__is_rare=True))
    return set(qs.values_list("species_id", flat=True))


@login_required
@require_GET
def comparatii_plante_data(request):
    """Return JSON summary for selected reserves. Params: res=1,2,3; rare=1.
    Response: { reserves: [ {id,name},...], common:[{id,name}], unique:{rid:[...]}, totalDistinct:N }
    """
    res_param = (request.GET.get("res") or "").strip()
    rare_flag = (request.GET.get("rare") or "1").strip() in ("1", "true", "yes", "on")
    ids = []
    if res_param:
        for part in res_param.split(','):
            part = part.strip()
            if part.isdigit():
                ids.append(int(part))
    ids = ids[:3]
    reserves = list(Reserve.objects.filter(id__in=ids).only("id", "name"))
    id_to_species = {}
    for r in reserves:
        id_to_species[r.id] = _species_ids_for_reserve(r.id, only_rare=rare_flag)
    # Compute sets
    common_ids = set.intersection(*id_to_species.values()) if id_to_species else set()
    all_ids = set().union(*id_to_species.values()) if id_to_species else set()
    unique = {}
    for rid, s in id_to_species.items():
        others = all_ids - s
        unique[rid] = list(s - (all_ids - s))  # unique to this reserve
    # Fetch species names
    species_map = { s.id: s.denumire_stiintifica for s in Species.objects.filter(id__in=all_ids).only("id","denumire_stiintifica") }
    def to_rows(id_list):
        return [ {"id": sid, "name": species_map.get(sid, str(sid))} for sid in sorted(id_list) ]
    data = {
        "reserves": [ {"id": r.id, "name": r.name} for r in reserves ],
        "common": to_rows(common_ids),
        "unique": { str(rid): to_rows(uids) for rid, uids in unique.items() },
        "totalDistinct": len(all_ids),
    }
    return JsonResponse(data)


@login_required
@require_GET
def comparatii_plante_export(request):
    kind = (request.GET.get("format") or request.GET.get("export") or "csv").lower()
    res_param = (request.GET.get("res") or "").strip()
    rare_flag = (request.GET.get("rare") or "1").strip() in ("1", "true", "yes", "on")
    ids = []
    if res_param:
        for part in res_param.split(','):
            if part.strip().isdigit():
                ids.append(int(part.strip()))
    ids = ids[:3]
    reserves = list(Reserve.objects.filter(id__in=ids).only("id", "name"))
    id_to_species = {}
    for r in reserves:
        id_to_species[r.id] = _species_ids_for_reserve(r.id, only_rare=rare_flag)
    common_ids = set.intersection(*id_to_species.values()) if id_to_species else set()
    all_ids = set().union(*id_to_species.values()) if id_to_species else set()
    species_map = { s.id: s.denumire_stiintifica for s in Species.objects.filter(id__in=all_ids).only("id","denumire_stiintifica") }

    headers = ["Tip", "Specie (științific)"] + [r.name for r in reserves]
    def rows_iter():
        yield headers
        # Common
        for sid in sorted(common_ids):
            yield ["Comun", species_map.get(sid, sid)] + ["Da"]*len(reserves)
        # Unique per reserve
        for r in reserves:
            only = id_to_species[r.id] - (all_ids - id_to_species[r.id])
            for sid in sorted(only):
                row = [f"Unic {r.name}", species_map.get(sid, sid)]
                for rr in reserves:
                    row.append("Da" if rr.id == r.id else "Nu")
                yield row

    # Filename
    base = "comparatie_plante_" + "_".join([r.name.replace(" ", "_") for r in reserves])
    if kind == "xlsx":
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            kind = "csv"
        else:
            wb = Workbook(); ws = wb.active; ws.title = "Comparatie";
            for row in rows_iter():
                ws.append(row)
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = f'attachment; filename="{base}.xlsx"'
            return resp

    return _stream_csv(f"{base}.csv", headers, rows_iter())


@login_required
@require_GET
def comparatii_plante_years(request):
    """Returns rare-only available years for one or multiple reserves.
    Params:
      - base: reserve id for base (optional)
      - res: comma-separated reserve ids
    Response: { yearsByReserve: { reserveId: [years...] } }
    """
    base = request.GET.get("base")
    res_param = (request.GET.get("res") or "").strip()
    ids = []
    if base and base.isdigit():
        ids.append(int(base))
    if res_param:
        for part in res_param.split(','):
            if part.strip().isdigit():
                ids.append(int(part.strip()))
    ids = list(dict.fromkeys(ids))  # dedupe preserve order
    yearsByReserve = {}
    for rid in ids:
        ys = (Occurrence.objects
              .filter(reserve_id=rid)
              .filter(Q(is_rare=True) | Q(species__is_rare=True))
              .exclude(year__isnull=True)
              .values_list("year", flat=True)
              .distinct())
        years = sorted(set(int(y) for y in ys if y), reverse=True)
        yearsByReserve[str(rid)] = years
    return JsonResponse({"yearsByReserve": yearsByReserve})


@login_required
def viz_situri_detail(request, pk: int):
    s = get_object_or_404(Site, pk=pk)
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    return render(request, "core/viz_situri_detail.html", {"s": s, "is_admin": is_admin})


@login_required
@require_POST
def update_site_meta(request, pk: int):
    is_admin = request.user.is_staff or request.user.groups.filter(name__iexact="Administrators").exists()
    if not is_admin:
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)

    site = get_object_or_404(Site, pk=pk)

    def norm_text(v):
        return strip_tags((v or "").strip()) or None

    changed = {}
    # Editable fields similar to reserves (adapted to Site model)
    for field in ("name", "code"):
        if field in request.POST:
            val = norm_text(request.POST.get(field))
            if getattr(site, field) != val:
                setattr(site, field, val)
                changed[field] = val

    # Numeric surface
    if "surface_ha" in request.POST:
        raw = (request.POST.get("surface_ha") or "").strip()
        if raw != "":
            try:
                num = float(raw)
                if num < 0:
                    return JsonResponse({"ok": False, "error": "Area must be positive"}, status=400)
            except ValueError:
                return JsonResponse({"ok": False, "error": "Invalid area"}, status=400)
        else:
            num = None
        if site.surface_ha != num:
            site.surface_ha = num
            changed["surface_ha"] = num

    # Coordinates
    lat = request.POST.get("latitude")
    lon = request.POST.get("longitude")
    if lat is not None or lon is not None:
        def parse_or_none(x):
            x = (x or "").strip()
            return None if x == "" else float(x)
        try:
            lat_v = parse_or_none(lat)
            lon_v = parse_or_none(lon)
            if lat_v is not None and not (-90.0 <= lat_v <= 90.0):
                return JsonResponse({"ok": False, "error": "Latitude out of range"}, status=400)
            if lon_v is not None and not (-180.0 <= lon_v <= 180.0):
                return JsonResponse({"ok": False, "error": "Longitude out of range"}, status=400)
        except ValueError:
            return JsonResponse({"ok": False, "error": "Invalid coordinates"}, status=400)
        if site.latitude != lat_v:
            site.latitude = lat_v
            changed["latitude"] = lat_v
        if site.longitude != lon_v:
            site.longitude = lon_v
            changed["longitude"] = lon_v

    # Integers
    if "bird_species_count" in request.POST:
        raw = (request.POST.get("bird_species_count") or "").strip()
        if raw == "":
            val = None
        else:
            try:
                val = int(raw)
                if val < 0:
                    return JsonResponse({"ok": False, "error": "Bird species count must be >= 0"}, status=400)
            except ValueError:
                return JsonResponse({"ok": False, "error": "Invalid bird species count"}, status=400)
        if site.bird_species_count != val:
            site.bird_species_count = val
            changed["bird_species_count"] = val

    if "other_species_count" in request.POST:
        raw = (request.POST.get("other_species_count") or "").strip()
        if raw == "":
            val2 = None
        else:
            try:
                val2 = int(raw)
                if val2 < 0:
                    return JsonResponse({"ok": False, "error": "Other species count must be >= 0"}, status=400)
            except ValueError:
                return JsonResponse({"ok": False, "error": "Invalid other species count"}, status=400)
        if site.other_species_count != val2:
            site.other_species_count = val2
            changed["other_species_count"] = val2

    # Booleans for biogeographic flags
    for bfield in ("ste", "conj"):
        if bfield in request.POST:
            sval = (request.POST.get(bfield) or "").strip().lower()
            newb = sval in ("true", "1", "yes", "on")
            if getattr(site, bfield) != newb:
                setattr(site, bfield, newb)
                changed[bfield] = newb

    if not changed:
        return JsonResponse({"ok": True, "changed": {}}, status=200)

    site.save()
    return JsonResponse({"ok": True, "changed": changed})


@login_required
def _export_sitehab(qs, kind):
    headers = ["Site", "Habitat (RO)", "Habitat (EN)", "Cod habitat", "An", "Suprafață (ha)", "Notițe"]

    if kind == "csv":
        def rows_iter():
            for r in qs:
                yield [
                    r.site.name,
                    getattr(r.habitat, "name_romanian", "") or "",
                    getattr(r.habitat, "name_english", "") or "",
                    getattr(r.habitat, "code", "") or "",
                    r.year,
                    r.surface if r.surface is not None else "",
                    r.notes or "",
                ]
        return _stream_csv("site_habitats.csv", headers, rows_iter())

    # XLSX – import "lazy"
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError:
        return HttpResponse(
            "Export XLSX necesită pachetul 'openpyxl'. Rulează: pip install openpyxl",
            content_type="text/plain; charset=utf-8",
            status=500
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "SiteHabitats"
    ws.append(headers)
    for r in qs:
        ws.append([
            r.site.name,
            getattr(r.habitat, "name_romanian", "") or "",
            getattr(r.habitat, "name_english", "") or "",
            getattr(r.habitat, "code", "") or "",
            r.year,
            r.surface if r.surface is not None else "",
            r.notes or "",
        ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    resp = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="site_habitats.xlsx"'
    return resp
    headers = ["Site", "Habitat (RO)", "Habitat (EN)", "Cod habitat", "An", "Suprafață (ha)", "Notițe"]
    rows = [
        [
            r.site.name,
            getattr(r.habitat, "name_romanian", "") or "",
            getattr(r.habitat, "name_english", "") or "",
            getattr(r.habitat, "code", "") or "",
            r.year,
            r.surface if r.surface is not None else "",
            r.notes or "",
        ]
        for r in qs
    ]

    if kind == "csv":
        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="site_habitats.csv"'
        return resp

    wb = Workbook()
    ws = wb.active
    ws.title = "SiteHabitats"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    resp = HttpResponse(stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="site_habitats.xlsx"'
    return resp